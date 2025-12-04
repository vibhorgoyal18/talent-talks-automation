from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook


class ExcelDataLoader:
    """Lightweight helper that exposes Excel-based test data as dictionaries."""

    def __init__(self, file_path: str | Path) -> None:
        self._path = Path(file_path)
        if not self._path.exists():
            raise FileNotFoundError(f"Excel file not found at {self._path}")
        self._workbook = load_workbook(self._path)

    def rows(self, sheet_name: str) -> list[dict[str, Any]]:
        sheet = self._workbook[sheet_name]
        header = [cell.value for cell in next(sheet.iter_rows(max_row=1))]
        data: list[dict[str, Any]] = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            data.append({header[idx]: value for idx, value in enumerate(row)})
        return data

    def find_by_key(self, sheet_name: str, key_column: str, key_value: str) -> dict[str, Any]:
        for row in self.rows(sheet_name):
            if str(row.get(key_column)) == key_value:
                return row
        raise KeyError(f"No data found in {sheet_name} for {key_column}={key_value}")
